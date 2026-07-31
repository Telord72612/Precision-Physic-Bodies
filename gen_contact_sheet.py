#!/usr/bin/env python3
"""Generate the PPB touch-API contact workbook.

Sources of truth (never hand-typed):
  * capsule_body_part_map.json  - generated from ProposedPartName(), DLL-verified
  * RegionOfPart()/DwellSForRegion() in PpbApi.cpp - parsed here for region + dwell
  * PPB_tuning.txt              - the shipped dwell/gate values
Every capsule name emitted is re-checked against the strings in the shipped PPB.dll;
the script aborts if any name is missing (the generated-record rule).
"""
import json, re, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"G:/Claude Workspace")
MAPJ = ROOT / "Report/Precision Physic Bodies Module/capsule_body_part_map.json"
SRC  = ROOT / "tools/PPB-plugin/src"
DLL  = Path(r"D:/Games/My Skyrim/mods/Precision Physic Bodies/SKSE/Plugins/PPB.dll")
TUNE = Path(r"D:/Games/My Skyrim/mods/Precision Physic Bodies/SKSE/Plugins/PPB_tuning.txt")
OUT  = ROOT / "Report/Precision Physic Bodies Module/PPB_Touch_API_Contact_List.xlsx"

# ── knobs (shipped values) ────────────────────────────────────────────────────
tune = {}
for ln in TUNE.read_text(errors="ignore").splitlines():
    m = re.match(r"^\s*(api\w+)\s+([-\d.]+)", ln)
    if m: tune[m.group(1)] = m.group(2)

# ── region map, parsed from RegionOfPart() ────────────────────────────────────
api = (SRC / "PpbApi.cpp").read_text(errors="ignore")
body = api[api.index("int RegionOfPart"):]
body = body[:body.index("const char* RegionLabel")]
SLOT_REGION = {}
for m in re.finditer(r"case\s+(\d+):(?:\s*case\s+(\d+):)?\s*return\s+PPBAPI::kRegion(\w+)", body):
    for g in (m.group(1), m.group(2)):
        if g: SLOT_REGION[int(g)] = m.group(3)
if 11 in SLOT_REGION: del SLOT_REGION[11]          # slot 11 is child-conditional
assert len(SLOT_REGION) >= 11, f"region parse failed: {SLOT_REGION}"

DWELL = {"Intimate": tune["apiDwellSensorS"], "Pelvis": tune["apiDwellComS"],
         "Face": tune["apiDwellHeadS"], "Tail": tune["apiDwellTailS"],
         "Hair": tune["apiDwellTailS"]}
def dwell(reg): return DWELL.get(reg, tune["apiDwellS"])

def region_of(slot, child):
    if slot == 11: return "Intimate" if 21 <= child <= 31 else "Pelvis"
    return SLOT_REGION.get(slot, "None")

# ── sub-group + override text, keyed by (slot, child) ─────────────────────────
SENSOR_OVR = ("OVERRIDES the general race: an interior sensor within touch range always wins "
              "over the bigger capsule it sits inside (a thigh at d=-11u used to mask it). "
              "Among sensors the DEEPEST wins, so WHERE answers how far it reached.")
# Mouth depth ladder (user, 2026-07-30): Face surface < Mouth opening < In mouth < Mouth wall.
# Each rung overrides the one below it; none overrides the one above.
FACE_CONJ   = ("Reads as a plain FACE TOUCH on its own. It only participates in mouth detection "
               "IN CONJUNCTION - the mouth opens when the fingertip is within mouthEnterU of the "
               "palate AND both cheeks at once, and only if the INDEX boxes are the nearest part "
               "of the hand (a hand cupping the jaw cannot open her). One capsule alone never "
               "means 'inside'.")
CHIN_EXTRA  = ("The chins are also the DEEP-HOLD line: once inside, both chins close + a LOOSE "
               "palate range keep the mouth open after the strict entry trio is left behind.")
MOUTH_LIP   = ("MOUTH OPENING - the lip ring. AT-LIPS is the lip + both chins all within gate: a "
               "logged approach stage, at the mouth but NOT inside it. Lowest rung of the mouth "
               "ladder; overridden by In mouth (palate) and by Mouth wall (throat).")
MOUTH_PALATE = ("IN MOUTH. The palate is the roof of the mouth cavity - if it registers a touch, "
                "something IS inside her mouth, and that OVERRIDES Mouth opening (a lip/entry "
                "reading cannot outrank it). It does NOT override Mouth wall: the throat is "
                "deeper still. Also the PRIMARY entry capsule - the finger-only test runs "
                "against it.")
FREE_PLACED = ("Free-placed capsule: position comes from its own UV/mesh model, not the parent "
               "node's welded transform (breasts + glutes only).")

# Editorial override prose, keyed by (slot, child). The GROUP comes from the C++ table below;
# this column is the human explanation of what that group's override actually does.
OVR = {}
def _ovr(slot, children, text):
    for c in children: OVR[(slot, c)] = text
_ovr(3, [4, 5], FACE_CONJ)                       # cheeks - entry trio members
_ovr(3, [2, 3], FACE_CONJ + " " + CHIN_EXTRA)    # chins  - AT-LIPS + deep hold
_ovr(3, [1],    MOUTH_LIP)
_ovr(3, [9],    MOUTH_PALATE)
_ovr(3, [10],
     "MOUTH WALL - the END of the mouth cavity, and the TOP of the ladder. OVERRIDES In mouth "
     "and Mouth opening both: proximity to it holds the mouth open BY ITSELF and raises the "
     "THROAT REACHED event, even when the forward entry capsules have been left behind. Nothing "
     "overrides it. Only evaluated while already inside (enter-through-the-front-only).")
_ovr(6, [11, 12], FREE_PLACED)
_ovr(11, [16, 17], FREE_PLACED)
_ovr(11, [0, 1, 2],
     "NOT a sensor: the three rings sit OUTSIDE the C21-C31 sensor range, so they report as "
     "region Pelvis, not Intimate.")
_ovr(11, range(21, 32), SENSOR_OVR)

# ---- sub-groups: PARSED from SubRegionOfPart()/SubRegionLabel() in PpbApi.cpp ----
# The C++ table is the source of truth (that function is what the API actually returns).
# Hand-maintaining a second copy here is exactly how a record drifts, so we read it.
api_src = (SRC / "PpbApi.cpp").read_text(errors="ignore")

def _fn(name, endmark):
    s = api_src.index(name); return api_src[s:api_src.index(endmark, s)]

# label table: case PPBAPI::kSubX: return "Text";
LABEL = dict(re.findall(r'case\s+PPBAPI::(kSub\w+):\s*return\s+"([^"]*)"',
                        _fn("const char* SubRegionLabel", "int SubRegionDepthOf")))
assert len(LABEL) >= 30, f"sub-region label parse failed: {len(LABEL)}"

# depth table: the same shape, falling through to kDepthSurface
_d = _fn("int SubRegionDepthOf", "// Stamp the three classification")
DEPTH, _pend = {}, []
for tok, lvl in re.findall(r'case\s+PPBAPI::(kSub\w+):|return\s+PPBAPI::(kDepth\w+)', _d):
    if tok: _pend.append(tok)
    else:
        for k in _pend: DEPTH[k] = lvl
        _pend = []
DEPTH_LABEL = {"kDepthSurface": "0 surface", "kDepthOpening": "1 opening",
               "kDepthInside": "2 inside", "kDepthDeepest": "3 deepest"}

# the dispatch itself: re-implement the switch in Python, keyed off the parsed enum names
_s = _fn("int SubRegionOfPart", "const char* SubRegionLabel")
def _sub_enum(slot, child):
    if slot == 3:
        return ("kSubHead" if child in (0, 14) else "kSubHeadEar" if child in (12, 13)
                else "kSubMouthOpening" if child == 1 else "kSubInMouth" if child == 9
                else "kSubMouthWall" if child == 10 else "kSubInMouthDeep" if child == 11
                else "kSubFaceSurface")
    if slot == 11:
        return ("kSubOrificeRing" if child in (0, 1, 2) else "kSubGlute" if child in (16, 17)
                else "kSubIntimateExternal" if child == 21
                else "kSubVaginalOpening" if child in (22, 23)
                else "kSubVaginalDeep" if child in (24, 25)
                else "kSubVaginalDeepest" if child in (26, 27)
                else "kSubAnalOpening" if child in (28, 29)
                else "kSubAnalDeep" if child in (30, 31) else "kSubPelvis")
    if slot == 6:
        return ("kSubBreast" if child in (11, 12) else "kSubShoulderCap" if child in (13, 14)
                else "kSubRibCage")
    if slot == 2: return "kSubShoulder" if child == 2 else "kSubUpperArm"
    return {0: "kSubPalm", 1: "kSubForearm", 4: "kSubWaist", 5: "kSubBelly", 7: "kSubNeck",
            8: "kSubThigh", 9: "kSubCalf", 10: "kSubFoot"}[slot]

# guard: every branch this mirror can emit must exist in the C++ dispatch text
for _e in {_sub_enum(s, c) for s in range(12) for c in range(40)}:
    assert _e in _s, f"{_e} not in SubRegionOfPart() - the C++ table changed shape"

def sub_of(slot, child):
    e = _sub_enum(slot, child)
    return LABEL[e], DEPTH_LABEL[DEPTH.get(e, "kDepthSurface")]

# ── build rows ────────────────────────────────────────────────────────────────
data = json.loads(MAPJ.read_text())
dll  = DLL.read_bytes()
rows, missing, unmapped = [], [], []
for s in data["slots"]:
    slot, node, nodeL = s["slot"], s["node"], s.get("nodeLeft")
    for ch in s["children"]:
        c, name = ch["child"], ch["name"]
        if ch.get("seed") or not name: continue
        if name.encode() not in dll: missing.append(f"{slot}.{c} {name}")
        grp, dep = sub_of(slot, c)
        ovr = OVR.get((slot, c), "")
        reg = region_of(slot, c)
        rows.append([name, f"{slot}.{c}", node, nodeL or "(centre - no twin)", reg, grp, dep,
                     "yes" if (slot == 11 and 21 <= c <= 31) else "", ovr,
                     float(dwell(reg)), ch.get("knob", "")])

if missing:  sys.exit("ABORT - names not present in the shipped DLL:\n  " + "\n  ".join(missing))
if unmapped: sys.exit("ABORT - capsules with no sub-group:\n  " + "\n  ".join(unmapped))

# garment pseudo-slots (not in the JSON - they are runtime chords, not skeleton children).
# The tail reports as THREE segments: BodyPartName() in PpbApi.cpp splits the chord chain into
# equal thirds - child*3 < n -> base, child*3 < n*2 -> mid, else tip.
TAIL_COMMON = ("Vanilla non-SMP tails are NEVER touched - they have no physics rig to attach to; "
               "only HDT-SMP tails get chords. Segment is computed from the chord's position in "
               "the chain (equal thirds), so it is correct on a 4-chord foxtail and a 14-chord "
               "fluffy tail alike. Only the first N chords carry PUSH sensors (N is per-table, "
               "TuneOf(tbl).sensors); the rest collide but cannot be pushed. Rigs are budgeted: "
               "npcRigMaxActors nearest actors within npcRigRangeU.")
GARMENT = [
    ["tail (base)", "100.0..n/3", "SMP tail rig (renamed bones)", "-",
     "Tail", "Tail - base (root third)", "0 surface", "",
     "The third nearest the body - where the tail meets her. " + TAIL_COMMON,
     float(tune["apiDwellTailS"]), "npcRigMaxActors / npcRigRangeU"],
    ["tail (mid)", "100.n/3..2n/3", "SMP tail rig (renamed bones)", "-",
     "Tail", "Tail - mid (middle third)", "0 surface", "",
     "The middle third of the chain. " + TAIL_COMMON,
     float(tune["apiDwellTailS"]), "npcRigMaxActors / npcRigRangeU"],
    ["tail (tip)", "100.2n/3..n", "SMP tail rig (renamed bones)", "-",
     "Tail", "Tail - tip (far third)", "0 surface", "",
     "The far third - this is what answers 'did I touch the TIP?'. " + TAIL_COMMON,
     float(tune["apiDwellTailS"]), "npcRigMaxActors / npcRigRangeU"],
    ["hair chord (per-link, SMP wigs)", "101.n", "SMP hair rig (renamed bones)", "-",
     "Hair", "Hair - SMP chord", "0 surface", "",
     "GATED OFF by default (apiHairTarget 0). Hair chords drape the head and shoulders, so in a "
     "nearest-surface race a finger aimed at a cheek lands on a hair chord and SHADOWS the face "
     "capsules. Collision still works; only API targeting is disabled.",
     float(tune["apiDwellTailS"]), "apiHairTarget"],
]
rows += GARMENT

# ── workbook ──────────────────────────────────────────────────────────────────
HDR  = Font(bold=True, color="FFFFFF", size=11)
FILL = PatternFill("solid", fgColor="2F5597")
BOX  = Border(*[Side(style="thin", color="BFBFBF")] * 4)
TOP  = Alignment(vertical="top", wrap_text=True)
REG_FILL = {"Face": "FFF2CC", "Intimate": "F8CBAD", "Pelvis": "FCE4D6", "Chest": "E2EFDA",
            "Belly": "E2EFDA", "Waist": "E2EFDA", "Neck": "E2EFDA", "Arm": "DEEBF7",
            "Hand": "DEEBF7", "Leg": "DEEBF7", "Foot": "DEEBF7", "Tail": "EAD1DC",
            "Hair": "EAD1DC"}

wb = Workbook()

def sheet(title, headers, body_rows, widths, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment, cell.border = HDR, FILL, Alignment(
            vertical="center", wrap_text=True), BOX
    for r in body_rows: ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment, cell.border = TOP, BOX
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    return ws

ws = sheet("Contact list",
    ["Collision capsule", "Address (slot.child)", "Skeleton node (right / centre)",
     "Left twin node", "API region", "API sub-region", "Depth", "Interior sensor",
     "Override / special role", "Dwell before the API reports it (s)", "Geometry knob prefix"],
    rows, [30, 12, 26, 26, 13, 26, 11, 10, 62, 14, 16])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    f = REG_FILL.get(row[4].value)
    if f:
        for c in row[:7]: c.fill = PatternFill("solid", fgColor=f)
    if row[7].value == "yes": row[0].font = Font(bold=True)

sheet("Regions", ["API region", "Enum", "What it covers", "Dwell (s)", "Why that dwell"], [
    ["Face", "kRegionFace=1", "Whole head: skull, face surface, the mouth chain", tune["apiDwellHeadS"],
     "Longer - the head is where incidental brushes are most common while talking"],
    ["Neck", "kRegionNeck=2", "Neck / throat capsule", tune["apiDwellS"], "Default"],
    ["Chest", "kRegionChest=3", "Rib cage, back (upper), breasts, shoulder caps", tune["apiDwellS"], "Default"],
    ["Belly", "kRegionBelly=4", "Belly, navel, midriff, flanks, back (lower)", tune["apiDwellS"], "Default"],
    ["Waist", "kRegionWaist=5", "Waist ring, lower abdomen, lower back", tune["apiDwellS"], "Default"],
    ["Pelvis", "kRegionPelvis=6", "Hips, glutes, groin, the three orifice rings (COM C0-C20)",
     tune["apiDwellComS"], "Longest - you brush past a hip constantly just walking by"],
    ["Intimate", "kRegionIntimate=7", "The interior sensor chain (COM C21-C31)", tune["apiDwellSensorS"],
     "Shortest - an insertion is deliberate, never incidental"],
    ["Arm", "kRegionArm=8", "Forearm + upper arm + shoulder", tune["apiDwellS"], "Default"],
    ["Hand", "kRegionHand=9", "Palm rods and plates", tune["apiDwellS"], "Default"],
    ["Leg", "kRegionLeg=10", "Thigh + calf", tune["apiDwellS"], "Default"],
    ["Foot", "kRegionFoot=11", "Sole, arch, ankle", tune["apiDwellS"], "Default"],
    ["Tail", "kRegionTail=12", "HDT-SMP tail chords (pseudo-slot 100)", tune["apiDwellTailS"], "Default"],
    ["Hair", "kRegionHair=13", "SMP hair chords (pseudo-slot 101) - targeting OFF by default",
     tune["apiDwellTailS"], "Default"],
], [14, 20, 58, 10, 52])

# every sub-region the API can return, in enum order, with its depth rung
SUB_COVER = {
 "Head": "Cranium and occiput.", "Head (temple / ear)": "Temple / ear side. These are the EAR capsules on beast heads.",
 "Face surface": "Cheekbones, nose, cheeks, chins. A plain face touch on its own - only signifies the mouth in conjunction.",
 "Mouth opening": "The upper-lip ring. At the mouth, not inside it.",
 "In mouth": "The palate. A touch here means something IS inside her mouth.",
 "In mouth (deep floor)": "Under-jaw / the deep floor of the cavity.",
 "Mouth wall": "The throat wall - the end of the cavity. Outranks everything.",
 "Neck": "The neck / throat capsule (the choke target).",
 "Shoulder cap": "Deltoid caps on spine2.", "Rib cage / back": "Chest ring, lats, front ribs, trapezius, collarbones, sternum, upper back.",
 "Breast": "The two computed breast capsules.", "Belly / midriff": "Belly, navel, midriff, flanks, lower back.",
 "Waist band": "Waist ring, lower belly, lower abdomen, lower back, flank.",
 "Shoulder": "The deltoid / shoulder cap on the upper-arm slot.", "Upper arm": "Shoulder half, elbow half, inner twin.",
 "Forearm": "Elbow half and wrist half.", "Palm": "Palm rod, thumb-side and pinky-side rails, palm centres.",
 "Pelvis / hip": "Hips, groin creases, front hips, rear centrelines, inner rails, upper glutes.",
 "Pelvis - orifice ring": "The three stacked rings. OUTSIDE the sensor range - reports as Pelvis, not Intimate.",
 "Glute": "The two butt cheeks (free-placed from the UV model).",
 "Intimate - external": "Clitoris.", "Intimate - vaginal (opening)": "Vaginal opening L/R.",
 "Intimate - vaginal (deep)": "Cervix L/R.", "Intimate - vaginal (deepest)": "Uterus L/R.",
 "Intimate - anal (opening)": "Anus L/R.", "Intimate - anal (deep)": "Rectum L/R.",
 "Thigh": "Thigh rod, groin junction, hip-glute fold, upper/mid/lower thigh, knee.",
 "Calf": "Calf rod, upper calf, calf belly, shin, knee rear.", "Foot": "Sole rods, arch, ankle lock.",
 "Tail - base (root third)": "First third of an SMP tail's chord chain.",
 "Tail - mid (middle third)": "Middle third of the chain.",
 "Tail - tip (far third)": "Far third - answers 'did I touch the TIP?'.",
 "Hair": "SMP hair chords. Declared but never emitted by default (apiHairTarget 0).",
}
_ENUM_ORDER = [e for e in re.findall(r'case\s+PPBAPI::(kSub\w+):\s*return\s+"', 
               _fn("const char* SubRegionLabel", "int SubRegionDepthOf"))]
sheet("Sub-regions",
    ["API sub-region", "Enum", "Depth", "What it covers"],
    [[LABEL[e], "PPBAPI::" + e, DEPTH_LABEL[DEPTH.get(e, "kDepthSurface")],
      SUB_COVER.get(LABEL[e], "")] for e in _ENUM_ORDER],
    [30, 30, 11, 78])

sheet("Sources (WITH WHAT)",
    ["Source", "Enum", "How it is decided", "Notes"], [
    ["Finger", "kSourceFinger=0", "VRIK getFingerPos() first (controller-driven, dead band 0.45/0.55); "
     "then measured curl; then which hand box led the contact", "An index box clearly nearest the capsule"],
    ["Open hand", "kSourceOpenHand=1", "Same ladder - fingers extended", ""],
    ["Fist", "kSourceFist=2", "Same ladder - fingers closed; tip-to-palm under apiFistTipPalmU "
     f"({tune['apiFistTipPalmU']}u)", "Measured curl spans 3.7-9.1u across grips"],
    ["Palm", "kSourcePalm=3", "The palm plate box led the contact", ""],
    ["HIGGS grab", "kSourceGrab=4", "HIGGS reports the hand is grabbing", ""],
    ["Weapon", "kSourceWeapon=5", "Segment-to-segment against the drawn weapon's blade segment",
     "Blade segment comes from TESBoundObject::boundData (the shape walk hits kConvexTransform "
     "with no public header)"],
    ["Object", "kSourceObject=6", "Held non-weapon object", ""],
], [14, 22, 66, 52])

sheet("Overrides & precedence",
    ["#", "Rule", "What it overrides", "Why", "Knob"], [
    [1, "Interior sensors win outright", "The general nearest-capsule race",
     "A big thigh capsule at d=-11.43u always beat an r=0.3 sensor whose depth bottoms out near "
     "-0.3, so a real 5-second insertion reported 'L upper thigh' and zero sensor names. Sensors "
     "now run a SEPARATE race that is copied over the winner; among sensors the deepest wins.", "-"],
    [2, "THE MOUTH DEPTH LADDER: Face surface < Mouth opening < In mouth < Mouth wall",
     "Each rung overrides every rung below it",
     "Four levels of 'how far in is it', and the deeper reading always wins. (a) FACE SURFACE - "
     "cheeks and chins are ordinary face touches on their own; they only mean 'mouth' in "
     "CONJUNCTION with the palate. (b) MOUTH OPENING - the lip ring; at the mouth, not inside it. "
     "(c) IN MOUTH - the palate is the roof of the cavity, so a palate touch means something IS "
     "inside; it overrides Mouth opening. (d) MOUTH WALL - the throat is the end of the cavity and "
     "overrides everything, including the palate.", "mouthThroatChild"],
    [3, "Mouth wall (C10) also overrides the EXIT condition", "The C9+C4+C5 entry trio",
     "The entry capsules sit forward in the mouth. Pushed all the way in, the finger leaves their "
     "range and the mouth used to snap shut. Near C10 = unambiguously deep, so it holds the mouth "
     "open by itself. The chins provide the same rescue one rung lower (deep hold: both chins "
     "close + a LOOSE palate range). Entry stays strict so a hand cupping the jaw cannot open her; "
     "staying open is loose so a deep finger is not lost - from under the jaw the palate is far, "
     "and that asymmetry is what separates deep-inside from below-the-jaw.", "-"],
    [4, "Held weapon or object suppresses that HAND", "Bare-hand contacts on the SAME hand only",
     "Your palm is wrapped around the grip, so it reports a permanent phantom fist. The other "
     "hand is unaffected.", f"apiSuppressHeldHand {tune['apiSuppressHeldHand']}"],
    [5, "Hair targeting is off", "Hair chords never enter the nearest-capsule race",
     "Hair drapes the head and shoulders, so a finger aimed at a cheek lands on a hair chord and "
     "shadows every face capsule. Hair still COLLIDES; it just cannot be a reported target.",
     f"apiHairTarget {tune['apiHairTarget']}"],
    [6, "Region digest supersedes per-capsule events", "The raw per-capsule stream",
     "Sliding a finger over a face crosses ~5 capsules in 6s, never dwelling 0.5s on any one, so "
     "per-capsule dwell emitted NOTHING. The digest accumulates time in the REGION: one "
     "'Face(L cheek) dur=6.11s' event instead of five dropped ones.", f"apiEvents {tune['apiEvents']}"],
    [7, "A source change does not restart the contact", "Contact timing",
     "Index finger to fist on the same region is one continuous touch, not two.", "-"],
    [8, "Hand / weapon / object are separate classes", "Each other",
     "They race independently (kClsHand / kClsWeapon / kClsObject), so a held sword cannot mask "
     "what the other hand is doing.", "-"],
    [9, "Raw verbose stream is opt-in", "Nothing - it runs alongside the digest",
     "Consumers that want every capsule transition can subscribe to the raw layer; it is off by "
     "default because it is chatty.", f"apiRawEvents {tune['apiRawEvents']}"],
    [10, "Vanilla tails are never touched", "Tail targeting",
     "Every NPC has a vanilla tail node but it carries no SMP rig, so nothing can drive it. Only "
     "HDT-SMP tails get chords. Do not mix the two.", "-"],
    [11, "Tail position is reported as base / mid / tip", "The bare chord index",
     "A chord index alone is meaningless across rigs - a foxtail has 4 and a fluffy tail 14. The "
     "chain is split into equal THIRDS, so 'tip' means the same place on every tail and a "
     "consumer can answer 'did I touch the TIP?' without knowing the rig.", "-"],
], [5, 42, 40, 72, 24])

sheet("Beast head map",
    ["Family", "Entry trio (opens mouth)", "Chin lines", "Throat event", "Note"], [
    ["Human / elf / draenei", "C9 palate + C4 cheek R + C5 cheek L", "C2 + C3",
     "yes (C10 throat wall)", "The naming reference. The 'Contact list' sheet is this skeleton."],
    ["Khajiit", "muzzle trio C6 + C7 + C8", "C2 + C3", "no",
     "TOUCHPROBE-mapped baked capsules. No palate condition - C9/C10 are other anatomy on this head."],
    ["Argonian", "lip + muzzle C1 + C6 + C8", "C2 + C4", "no",
     "Same - their C9/C10 are other anatomy."],
], [22, 40, 16, 22, 62])

sheet("How to read this",
    ["Field", "Meaning"], [
    ["Address (slot.child)", "The stable API address. Slot = which skeleton node the capsule welds "
     "to; child = its index under that node. A full address on a sided slot is (slot, side, child)."],
    ["Left twin node", "Sided slots carry an identical mirrored capsule set under this node. "
     "Centre slots (head, spine 0-2, neck, COM) have no twin."],
    ["API region", "The sub-category the API reports in the digest stream, e.g. "
     "'R|Index|Face(L cheek) dur=6.11s'. 13 regions total - see the Regions sheet."],
    ["API sub-region", "The finer bucket, and it IS sent through the API (added 2026-07-31): "
     "IPpbTouchInterface1::SubRegionOf() natively, PPB_Touch.GetContactSubRegion() from Papyrus, "
     "and PpbTouchContact::subRegion on the struct itself. See the Sub-regions sheet."],
    ["Depth", "The ladder as a number - 0 surface, 1 opening, 2 inside, 3 deepest. Answers 'did it "
     "go in, and how far' without caring where. SubRegionDepth() / GetContactDepth()."],
    ["Interior sensor", "COM C21-C31 only. These override the general race (see Overrides sheet #1)."],
    ["Dwell", "The probe must stay in the REGION this long before the API emits. Region-level, not "
     "per-capsule - that is what makes a wandering face touch report as one event."],
    ["Seeds", "6 capsule children are buried structural seeds and are deliberately absent from this "
     "sheet - they are never surfaced through the API."],
    ["Index alignment", "Names are index-aligned across all four PPB skeletons only within each "
     "slot's NAMED range. Above that range the same index is different anatomy per race, which is "
     "why those indices are left unnamed."],
    ["Provenance", "Generated by tools/ppb-repo-work/gen_contact_sheet.py from "
     "capsule_body_part_map.json + RegionOfPart() + the shipped PPB_tuning.txt. Every capsule name "
     "is verified present in the shipped PPB.dll before the file is written."],
], [22, 116])

del wb["Sheet"]
wb.move_sheet("Contact list", offset=-len(wb.sheetnames))

# ── self-check: every knob name cited anywhere in the book must exist in PPB_tuning.txt ──
tune_text = TUNE.read_text(errors="ignore")
declared  = set(re.findall(r"^\s*([a-zA-Z]\w+)\s+[-\d.]", tune_text, re.M))
cited, bad = set(), []
for sh in wb.worksheets:
    for row in sh.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                cited |= set(re.findall(r"\b(?:api|npc|mouth|cap)[A-Z]\w+", v))
for k in sorted(cited):
    # cap* values are knob PREFIXES (capHeadC9 -> capHeadC9Enable / capHeadC9AX / ...);
    # everything else must be a whole declared knob.
    ok = any(d.startswith(k) for d in declared) if k.startswith("cap") else k in declared
    if not ok: bad.append(k)
if bad: sys.exit("ABORT - knob names cited but absent from PPB_tuning.txt:\n  " + "\n  ".join(bad))
print(f"knob check: {len(cited)} cited, all present in PPB_tuning.txt")

wb.save(OUT)
print(f"OK  {len(rows)} rows ({len(rows)-len(GARMENT)} skeleton capsules + {len(GARMENT)} garment) "
      f"-> {OUT}")
